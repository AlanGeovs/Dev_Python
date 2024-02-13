import os
import re
import email
import pandas as pd
from email.utils import parseaddr
from email.policy import default
from openpyxl import load_workbook

# Función para extraer la dirección IP del encabezado 'Received'
def extract_ip_from_received_headers(received_headers):
    ip_regex = r'[0-9]+(?:\.[0-9]+){3}'
    for header in reversed(received_headers):
        ips = re.findall(ip_regex, header)
        if ips:
            return ips[-1]
    return None

# Función para extraer solo la dirección de correo electrónico
def extract_email_address(raw_email):
    email_address = parseaddr(raw_email)[1]
    return email_address

# Función para procesar y extraer información del archivo .eml
def parse_eml(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        try:
            msg = email.message_from_file(file, policy=default)
        except Exception as e:
            print(f"No se pudo procesar el archivo {file_path}: {e}")
            return None
    
    remitente_nombre, remitente_email = parseaddr(msg.get('From'))
    destinatario_email = extract_email_address(msg.get('To'))
    otros_destinatarios_emails = [extract_email_address(addr) for addr in msg.get_all('Cc', [])]
    fecha = msg.get('Date')
    asunto = msg.get('Subject')
    received_headers = msg.get_all('Received')
    ip = extract_ip_from_received_headers(received_headers) if received_headers else None
    
    return {
        'nombre_remitente': remitente_nombre,
        'remitente': remitente_email,
        'destinatario': destinatario_email,
        'otros_destinatarios': ", ".join(otros_destinatarios_emails),
        'fecha': fecha,
        'asunto': asunto,
        'ip': ip
    }

# Función para procesar todos los archivos en un directorio
def process_eml_files(directory):
    data = []
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        parsed_email = parse_eml(file_path)
        if parsed_email:
            data.append(parsed_email)
    return data

# Procesar archivos y cargar o crear el archivo Excel
eml_directory = 'mails'  # Asegúrate de actualizar esta ruta
eml_data = process_eml_files(eml_directory)
excel_path = 'bd_correos.xlsx'  # Asegúrate de actualizar esta ruta

if os.path.exists(excel_path):
    workbook = load_workbook(filename=excel_path)
    sheet = workbook.active
else:
    df_new = pd.DataFrame(columns=["id", "nombre_remitente", "remitente", "destinatario", "otros_destinatarios", "fecha", "asunto", "ip"])
    df_new.to_excel(excel_path, index=False)
    workbook = load_workbook(filename=excel_path)
    sheet = workbook.active

last_id = sheet.cell(row=sheet.max_row, column=1).value if sheet.max_row > 1 else 0
last_id = int(last_id) if last_id is not None else 0

for email_data in eml_data:
    if email_data:
        last_id += 1
        sheet.append([
            last_id,
            email_data.get('nombre_remitente'),
            email_data.get('remitente'),
            email_data.get('destinatario'),
            email_data.get('otros_destinatarios'),
            email_data.get('fecha'),
            email_data.get('asunto'),
            email_data.get('ip')
        ])

workbook.save(filename=excel_path)
print(f"Archivo Excel actualizado: {excel_path}")
